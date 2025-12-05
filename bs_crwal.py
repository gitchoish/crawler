from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import pandas as pd
import time
from datetime import datetime
import os
import re

class NaverSmartStoreReviewCrawler:
    def __init__(self, product_url, rating_filter=None):
        """
        product_url: 제품 URL
        rating_filter: 평점 필터 
            - None: 모든 평점 (평점 없는 것 제외)
            - [5]: 5점만
            - [4, 5]: 4점과 5점만
            - [1, 2, 3]: 1~3점 (낮은 평점)
        """
        self.product_url = product_url
        self.rating_filter = rating_filter
        self.reviews = []
        self.driver = None
        
    def setup_driver(self):
        """Chrome 드라이버를 설정합니다."""
        print("=" * 60)
        print("🚀 네이버 스마트스토어 리뷰 크롤러 (평점 필터링)")
        print("=" * 60)
        
        if self.rating_filter:
            print(f"⭐ 평점 필터: {self.rating_filter}점만 수집")
        else:
            print("⭐ 평점 필터: 평점 있는 모든 리뷰 수집")
        
        print("\n[1/5] Chrome 드라이버 초기화 중...")
        
        options = webdriver.ChromeOptions()
        options.add_argument('--headless=new')
        options.add_argument('--disable-blink-features=AutomationControlled')
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        options.add_argument('user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
        options.add_argument('--no-sandbox')
        options.add_argument('--disable-dev-shm-usage')
        options.add_argument('--disable-gpu')
        options.add_argument('--window-size=1920,1080')
        options.add_argument('--disable-extensions')
        options.add_argument('--disable-images')
        options.add_argument('--blink-settings=imagesEnabled=false')
        
        try:
            # 방법 1: Selenium 4의 자동 드라이버 관리 사용
            self.driver = webdriver.Chrome(options=options)
        except Exception as e1:
            print(f"   ⚠️ 자동 드라이버 실패, webdriver-manager 시도...")
            try:
                # 방법 2: webdriver-manager 사용
                from webdriver_manager.chrome import ChromeDriverManager
                from selenium.webdriver.chrome.service import Service
                service = Service(ChromeDriverManager().install())
                self.driver = webdriver.Chrome(service=service, options=options)
            except Exception as e2:
                print(f"   ❌ 드라이버 초기화 실패: {e2}")
                raise
        
        self.driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        
        print("✅ 드라이버 초기화 완료\n")
        
    def navigate_to_product(self):
        """제품 페이지로 이동합니다."""
        print(f"[2/5] 제품 페이지 로딩 중...")
        self.driver.get(self.product_url)
        time.sleep(2)
        print("✅ 페이지 로딩 완료\n")
        
    def click_review_tab(self):
        """리뷰 탭을 클릭합니다."""
        print("[3/5] 리뷰 탭으로 이동 중...")
        
        try:
            wait = WebDriverWait(self.driver, 10)
            review_tab = wait.until(
                EC.element_to_be_clickable((By.CSS_SELECTOR, "a[href='#REVIEW']"))
            )
            self.driver.execute_script("arguments[0].click();", review_tab)
            time.sleep(3)
            print("✅ 리뷰 탭 이동 완료\n")
        except Exception as e:
            print(f"⚠️  리뷰 탭 클릭 실패: {e}")
            time.sleep(3)
    
    def get_first_review_signature(self):
        """첫 번째 리뷰의 시그니처를 반환합니다."""
        try:
            review_list = self.driver.find_element(By.CSS_SELECTOR, "#REVIEW ul")
            first_review = review_list.find_element(By.CSS_SELECTOR, "li:first-child")
            signature = first_review.text[:100]
            return signature
        except:
            return None
    
    def is_rating_match(self, rating):
        """평점이 필터 조건에 맞는지 확인합니다."""
        if not rating:
            return False
        
        try:
            rating_num = int(rating)
            
            if self.rating_filter is None:
                return True
            
            return rating_num in self.rating_filter
        except:
            return False
    
    def extract_reviews_from_current_page(self):
        """현재 페이지에서 리뷰를 추출합니다."""
        page_reviews = []
        
        try:
            # 스크롤하여 리뷰 로딩
            self.driver.execute_script("window.scrollTo(0, 1500);")
            time.sleep(3)
            
            # 리뷰 리스트 찾기
            review_list = self.driver.find_element(By.CSS_SELECTOR, "#REVIEW div.JHZoCyHfg7 div.HTT4L8U0CU ul")
            review_items = review_list.find_elements(By.CSS_SELECTOR, "li")
            
            for idx, item in enumerate(review_items):
                try:
                    review_data = {
                        'content': '',
                        'rating': '',
                        'date': '',
                        'reviewer': '',
                        'has_photo': False,
                        'tags': ''
                    }
                    
                    full_text = item.text.strip()
                    
                    if not full_text or len(full_text) < 5:
                        continue
                    
                    # 평점 추출
                    rating_match = re.search(r'평점\s*(\d+)', full_text)
                    if rating_match:
                        review_data['rating'] = rating_match.group(1)
                    
                    # 평점 필터링 체크
                    if not self.is_rating_match(review_data['rating']):
                        continue
                    
                    # 리뷰 본문 추출
                    content_found = False
                    
                    try:
                        content_div = item.find_element(By.CSS_SELECTOR, "div.HakaEZ240l")
                        content_text = content_div.text.strip()
                        if content_text and len(content_text) > 3:
                            review_data['content'] = content_text
                            content_found = True
                    except:
                        pass
                    
                    if not content_found:
                        try:
                            content_div = item.find_element(By.CSS_SELECTOR, "div.O2M37e85_1 div.HakaEZ240l")
                            content_text = content_div.text.strip()
                            if content_text and len(content_text) > 3:
                                review_data['content'] = content_text
                                content_found = True
                        except:
                            pass
                    
                    if not content_found:
                        try:
                            content_div = item.find_element(By.CSS_SELECTOR, "div.IwcuBUIAKf div[class*='HakaEZ']")
                            content_text = content_div.text.strip()
                            if content_text and len(content_text) > 3:
                                review_data['content'] = content_text
                                content_found = True
                        except:
                            pass
                    
                    if not content_found:
                        lines = full_text.split('\n')
                        content_lines = []
                        
                        for line in lines:
                            line = line.strip()
                            if len(line) > 5:
                                if not re.match(r'^평점\d+$', line):
                                    if not re.match(r'^\d{2}\.\d{2}\.\d{2}$', line):
                                        if '신고' not in line:
                                            if not re.match(r'^[a-z0-9*]+$', line):
                                                content_lines.append(line)
                        
                        if content_lines:
                            review_data['content'] = ' '.join(content_lines)
                            content_found = True
                    
                    if not content_found or len(review_data['content']) < 3:
                        continue
                    
                    # 날짜 추출
                    date_match = re.search(r'(\d{2}\.\d{2}\.\d{2})', full_text)
                    if date_match:
                        review_data['date'] = date_match.group(1)
                    
                    # 작성자 추출
                    reviewer_match = re.search(r'([a-z0-9*]+)\s*\d{2}\.\d{2}\.\d{2}', full_text)
                    if reviewer_match:
                        review_data['reviewer'] = reviewer_match.group(1)
                    
                    # 태그 추출
                    tags = []
                    tag_keywords = ['유통기한', '포장', '편리', '배송', '한달사용', '재구매', '가성비']
                    for keyword in tag_keywords:
                        if keyword in full_text:
                            tags.append(keyword)
                    review_data['tags'] = ', '.join(tags) if tags else ''
                    
                    # 사진 리뷰 여부
                    try:
                        imgs = item.find_elements(By.TAG_NAME, "img")
                        review_data['has_photo'] = len(imgs) > 1
                    except:
                        pass
                    
                    page_reviews.append(review_data)
                    
                except Exception as e:
                    continue
            
        except Exception as e:
            print(f"   ⚠️  페이지 추출 오류: {e}")
        
        return page_reviews
    
    def click_next_button(self):
        """'다음' 버튼을 클릭하여 다음 페이지 그룹으로 이동합니다."""
        try:
            # 페이지네이션 영역으로 스크롤
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(1)
            
            # '다음' 버튼 찾기 (nth-child로)
            next_button = None
            
            try:
                # '다음' 버튼은 보통 마지막에 위치
                next_button = self.driver.find_element(By.CSS_SELECTOR, 
                    "#REVIEW div.JHZoCyHfg7 div.HTT4L8U0CU > div > div > a.JY2WGJ4hXh.I3i1NSoFdB")
            except:
                pass
            
            if not next_button:
                try:
                    next_button = self.driver.find_element(By.CSS_SELECTOR, 
                        "#REVIEW div.HTT4L8U0CU > div > div > a[class*='I3i1NSoFdB']")
                except:
                    pass
            
            if not next_button or not next_button.is_displayed():
                return False
            
            # 버튼으로 스크롤
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_button)
            time.sleep(0.5)
            
            # 클릭
            self.driver.execute_script("arguments[0].click();", next_button)
            time.sleep(4)
            
            # 리뷰 영역으로 다시 스크롤
            self.driver.execute_script("window.scrollTo(0, 1500);")
            time.sleep(3)
            
            return True
            
        except Exception as e:
            return False
    
    def click_page_by_nth_child(self, nth_child):
        """nth-child 선택자로 페이지를 클릭합니다."""
        try:
            old_signature = self.get_first_review_signature()
            
            # 페이지네이션 영역으로 스크롤
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(0.5)
            
            # nth-child CSS 선택자로 페이지 버튼 찾기
            page_button = None
            
            try:
                selector = f"#REVIEW div.JHZoCyHfg7 div.HTT4L8U0CU > div > div > a:nth-child({nth_child})"
                page_button = self.driver.find_element(By.CSS_SELECTOR, selector)
            except Exception as e:
                return False
            
            if not page_button or not page_button.is_displayed():
                return False
            
            # 버튼으로 스크롤
            self.driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", page_button)
            time.sleep(0.5)
            
            # 클릭
            self.driver.execute_script("arguments[0].click();", page_button)
            time.sleep(4)
            
            # 리뷰 영역으로 스크롤
            self.driver.execute_script("window.scrollTo(0, 1500);")
            time.sleep(3)
            
            # 페이지가 실제로 변경되었는지 확인
            new_signature = self.get_first_review_signature()
            
            if old_signature and new_signature:
                if old_signature == new_signature:
                    return False
            
            return True
            
        except Exception as e:
            return False
    
    def collect_reviews_by_pagination(self, max_reviews=1000):
        """페이지네이션을 통해 리뷰를 수집합니다."""
        print(f"[4/5] 페이지별 리뷰 수집 중... (목표: {max_reviews}개)")
        print("=" * 60)
        
        current_page = 1
        max_pages = 100
        total_collected = 0
        start_time = time.time()
        
        seen_reviews = set()
        consecutive_failures = 0
        
        # nth-child 매핑
        # 페이지 1~10: nth-child(2)~(11)
        # 페이지 11: '다음' 버튼 클릭
        # 페이지 12~20: nth-child(3)~(11)
        # 페이지 21: '다음' 버튼 클릭
        # 페이지 22~30: nth-child(3)~(11)
        
        while total_collected < max_reviews and current_page <= max_pages:
            page_reviews = self.extract_reviews_from_current_page()
            
            new_count = 0
            duplicate_count = 0
            
            if page_reviews:
                for review in page_reviews:
                    if total_collected >= max_reviews:
                        break
                    
                    review_key = f"{review.get('date', '')}_{review.get('reviewer', '')}_{review['content'][:80]}"
                    
                    if review_key in seen_reviews:
                        duplicate_count += 1
                        continue
                    
                    seen_reviews.add(review_key)
                    review['number'] = total_collected + 1
                    self.reviews.append(review)
                    total_collected += 1
                    new_count += 1
                
                elapsed = int(time.time() - start_time)
                
                if duplicate_count > 0:
                    print(f"📄 페이지 {current_page}: {new_count}개 수집 ({duplicate_count}개 중복) | 누적: {total_collected}개 | {elapsed}초")
                else:
                    print(f"📄 페이지 {current_page}: {new_count}개 수집 | 누적: {total_collected}개 | {elapsed}초")
                
                if new_count == 0:
                    consecutive_failures += 1
                else:
                    consecutive_failures = 0
                
                if consecutive_failures >= 5:
                    print(f"\n⚠️  5페이지 연속 수집 실패. 종료합니다.")
                    break
            else:
                consecutive_failures += 1
                print(f"📄 페이지 {current_page}: 0개 수집 | 누적: {total_collected}개")
                
                if consecutive_failures >= 5:
                    break
            
            if total_collected >= max_reviews:
                print(f"\n✅ 목표 달성! {total_collected}개 수집 ({int(time.time() - start_time)}초)")
                break
            
            # 다음 페이지로 이동
            current_page += 1
            
            # nth-child 값 계산
            if current_page <= 10:
                # 페이지 1~10: nth-child(2)~(11)
                nth_child = current_page + 1
                
                if not self.click_page_by_nth_child(nth_child):
                    consecutive_failures += 1
                    if consecutive_failures >= 3:
                        break
                        
            elif current_page % 10 == 1:
                # 페이지 11, 21, 31... : '다음' 버튼 클릭
                print(f"\n   ⏭️  페이지 그룹 전환 중... (페이지 {current_page}로)")
                
                if self.click_next_button():
                    print(f"   ✓ '다음' 버튼 클릭 완료\n")
                else:
                    print(f"   ⚠️  '다음' 버튼을 찾을 수 없습니다. 수집 종료.")
                    break
                    
            else:
                # 페이지 12~20, 22~30...: nth-child(3)~(11)
                page_in_group = ((current_page - 1) % 10) + 1
                nth_child = page_in_group + 1
                
                if not self.click_page_by_nth_child(nth_child):
                    consecutive_failures += 1
                    if consecutive_failures >= 3:
                        break
        
        print(f"\n✅ 리뷰 수집 완료 (총 {len(self.reviews)}개)\n")
    
    def save_to_csv(self, filename=None):
        if not self.reviews:
            return None
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"denps_reviews_{timestamp}.csv"
        
        df = pd.DataFrame(self.reviews)
        column_order = ['number', 'date', 'rating', 'reviewer', 'content', 'tags', 'has_photo']
        df = df[[col for col in column_order if col in df.columns]]
        
        df.to_csv(filename, index=False, encoding='utf-8-sig')
        print(f"💾 CSV 저장: {filename}")
        
        return filename
    
    def save_to_excel(self, filename=None):
        if not self.reviews:
            return None
        
        if filename is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"denps_reviews_{timestamp}.xlsx"
        
        df = pd.DataFrame(self.reviews)
        
        column_mapping = {
            'number': '번호',
            'date': '작성일',
            'rating': '평점',
            'reviewer': '작성자',
            'content': '리뷰내용',
            'tags': '태그',
            'has_photo': '사진리뷰'
        }
        
        df = df.rename(columns={k: v for k, v in column_mapping.items() if k in df.columns})
        
        if '사진리뷰' in df.columns:
            df['사진리뷰'] = df['사진리뷰'].apply(lambda x: 'O' if x else 'X')
        
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='리뷰데이터')
            
            worksheet = writer.sheets['리뷰데이터']
            
            column_widths = {
                '번호': 8,
                '작성일': 12,
                '평점': 8,
                '작성자': 15,
                '리뷰내용': 80,
                '태그': 30,
                '사진리뷰': 10
            }
            
            for col_name, width in column_widths.items():
                if col_name in df.columns:
                    col_idx = list(df.columns).index(col_name) + 1
                    col_letter = worksheet.cell(row=1, column=col_idx).column_letter
                    worksheet.column_dimensions[col_letter].width = width
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1):
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
                    
                    if cell.column_letter in ['A', 'C', 'G']:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        print(f"💾 Excel 저장: {filename}")
        print(f"   📂 위치: {os.path.abspath(filename)}")
        
        return filename
    
    def print_summary(self):
        if not self.reviews:
            print("❌ 수집된 리뷰가 없습니다.")
            return
        
        df = pd.DataFrame(self.reviews)
        
        print("\n" + "=" * 60)
        print("📊 리뷰 수집 결과 요약")
        print("=" * 60)
        print(f"\n✅ 총 수집: {len(self.reviews)}개")
        
        if self.rating_filter:
            print(f"⭐ 필터 적용: {self.rating_filter}점만 수집")
        
        if df['rating'].notna().any() and df['rating'].ne('').any():
            df['rating_num'] = pd.to_numeric(df['rating'], errors='coerce')
            avg_rating = df['rating_num'].mean()
            if not pd.isna(avg_rating):
                print(f"⭐ 평균 평점: {avg_rating:.2f}점")
            
            print(f"\n평점 분포:")
            rating_counts = df['rating_num'].value_counts().sort_index(ascending=False)
            for rating, count in rating_counts.items():
                if not pd.isna(rating):
                    print(f"  {int(rating)}점: {count}개")
        
        photo_count = df['has_photo'].sum()
        print(f"\n📷 사진 리뷰: {photo_count}개 ({photo_count/len(self.reviews)*100:.1f}%)")
        
        df['content_length'] = df['content'].str.len()
        avg_length = df['content_length'].mean()
        print(f"📝 평균 리뷰 길이: {avg_length:.0f}자")
        
        print("\n📝 샘플 (상위 3개):")
        print("-" * 60)
        for i, review in enumerate(self.reviews[:3], 1):
            print(f"\n[{i}] 평점: {review['rating']}점 | 날짜: {review.get('date', '')}")
            content = review['content'][:100] + "..." if len(review['content']) > 100 else review['content']
            print(f"    {content}")
    
    def close(self):
        if self.driver:
            self.driver.quit()
    
    def run(self, max_reviews=1000):
        try:
            self.setup_driver()
            self.navigate_to_product()
            self.click_review_tab()
            self.collect_reviews_by_pagination(max_reviews)
            
            if self.reviews:
                self.print_summary()
                csv_file = self.save_to_csv()
                excel_file = self.save_to_excel()
                
                print("\n" + "=" * 60)
                print("✅ 크롤링 완료!")
                print("=" * 60)
            else:
                print("\n❌ 조건에 맞는 리뷰를 찾지 못했습니다.")
            
            return self.reviews
            
        except Exception as e:
            print(f"\n❌ 오류: {e}")
            import traceback
            traceback.print_exc()
            return []
        finally:
            self.close()
            print("\n프로그램을 종료합니다.")


# ============================================================
# 실행 코드
# ============================================================
if __name__ == "__main__":
    PRODUCT_URL = "https://brand.naver.com/denps/products/11261507716"
    MAX_REVIEWS = 1000
    
    # 모든 평점 수집
    RATING_FILTER = None
    
    print("\n" + "⭐ " * 20)
    print("네이버 스마트스토어 리뷰 크롤러 (nth-child 방식)")
    print("⭐ " * 20)
    print(f"📂 저장 위치: {os.getcwd()}")
    
    if RATING_FILTER:
        print(f"⭐ 수집 평점: {RATING_FILTER}점만")
    else:
        print("⭐ 수집 평점: 평점 있는 모든 리뷰")
    print()
    
    crawler = NaverSmartStoreReviewCrawler(PRODUCT_URL, rating_filter=RATING_FILTER)
    reviews = crawler.run(max_reviews=MAX_REVIEWS)
